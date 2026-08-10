from __future__ import annotations

import base64
from calendar import month_name
from datetime import timedelta
from functools import wraps
from io import BytesIO
from pathlib import Path

import qrcode

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.core.exceptions import PermissionDenied
from django.core.exceptions import ValidationError
from django.core.signing import BadSignature, Signer
from django.db.models import Count, Prefetch, Q
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.templatetags.static import static
from django.urls import reverse
from django.utils import timezone
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from almacen_app.models import Institucion

from .forms import (
    AnioForm,
    AsociacionForm,
    AsociacionUsuarioForm,
    ChecklistAnioItemFormSet,
    DepartamentoConstanciaForm,
    ExpedienteCAIMUSForm,
    FirmaConstanciaForm,
    ItemChecklistFormSet,
    RevisorConstanciaForm,
    RevisionExpedienteForm,
)
from .models import (
    Anio,
    Asociacion,
    AsociacionUsuario,
    DepartamentoConstancia,
    EntradaRevisionAdmin,
    ExpedienteCAIMUS,
    ExpedienteEstadoHistorial,
    InformeEstadoHistorial,
    InformeMensual,
    HistorialItemExpediente,
    FirmaConstancia,
    ItemChecklistCAIMUS,
    ConfiguracionInformeAnio,
    NotificacionAdmin,
    NotificacionAsociacion,
    ResolucionInformeMensual,
    ResolucionExpediente,
    RevisorConstancia,
    crear_notificacion_asociacion,
    crear_notificacion_admin,
    crear_entrada_revision_admin,
    crear_items_expediente,
    crear_informes_mensuales,
    asegurar_configuracion_informes_anio,
    informe_mes_requerido,
    generar_correlativo,
    generar_correlativo_informe,
    obtener_configuracion_informes_anio,
    resumen_informes_asociacion,
)
from .mixins import admin_required, asociacion_required
from .permissions import (
    expediente_esta_completo,
    expediente_items_100_aprobados,
    get_asociaciones_usuario,
    is_admin,
    is_asociacion,
    is_informatica,
    user_can_download_resolucion,
    user_has_asociacion_access,
    user_has_expediente_access,
)

ALLOWED_EXPEDIENTE_ITEM_EXTENSIONS = {".pdf", ".xls", ".xlsx"}
ALLOWED_EXPEDIENTE_ITEM_CONTENT_TYPES = {
    "application/pdf",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}
from .utils import obtener_entradas_bandeja_admin, resumen_dashboard_admin


PUBLIC_MONTHS = (
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
)


def _estado_publico_expediente(expediente, aprobados, total, correlativo):
    """Return the truthful public state; approval always requires every active item."""
    if expediente is None:
        return "sin_iniciar", "Sin iniciar"
    if expediente.estado == ExpedienteCAIMUS.ESTADO_APROBADO and total and aprobados == total and correlativo:
        return "aprobado", "Aprobado"
    labels = {
        ExpedienteCAIMUS.ESTADO_EN_REVISION: ("revision", "En revisión"),
        ExpedienteCAIMUS.ESTADO_RECHAZADO: ("rechazado", "Rechazado / Con observaciones"),
        ExpedienteCAIMUS.ESTADO_APROBADO: ("revision", "En revisión"),
    }
    return labels.get(expediente.estado, ("borrador", "Borrador"))


def _datos_publicos_asociacion(asociacion, detalle=False):
    expediente = getattr(asociacion, "expedientes_publicos", None)
    items = list(getattr(expediente, "items_publicos", [])) if expediente else []
    aprobados = sum(item.estado_item == ItemChecklistCAIMUS.ESTADO_APROBADO for item in items)
    total = len(items)
    porcentaje = int(aprobados * 100 / total) if total else 0
    resolucion = getattr(expediente, "resolucion_publica", None) if expediente else None
    correlativo = resolucion.correlativo if resolucion and resolucion.correlativo else None
    estado_clave, estado = _estado_publico_expediente(expediente, aprobados, total, correlativo)
    if estado_clave != "aprobado":
        correlativo = None

    configuraciones = list(getattr(asociacion.anio, "config_publica", []))
    requeridos = {config.mes for config in configuraciones if config.requerido}
    no_requeridos = {config.mes for config in configuraciones if not config.requerido}
    # An unconfigured active year defaults to requiring all months, consistently
    # with informe_mes_requerido().
    if not configuraciones:
        requeridos = set(range(1, 13))
    informes = {informe.mes: informe for informe in getattr(asociacion, "informes_publicos", [])}
    informes_aprobados = sum(
        mes in informes and informes[mes].estado == InformeMensual.ESTADO_APROBADO
        for mes in requeridos
    )
    informes_revision = sum(
        mes in informes and informes[mes].estado == InformeMensual.ESTADO_EN_REVISION
        for mes in requeridos
    )
    informes_pendientes = len(requeridos) - informes_aprobados - informes_revision
    fechas = []
    if expediente:
        fechas.append(expediente.actualizado_en)
        fechas.extend(item.fecha_actualizacion for item in items if item.fecha_actualizacion)
    fechas.extend(informe.actualizado_en for informe in informes.values() if informe.actualizado_en)

    data = {
        "pk": asociacion.pk,
        "nombre": asociacion.nombre,
        "codigo": asociacion.codigo,
        "anio": asociacion.anio.anio,
        "estado_clave": estado_clave,
        "estado": estado,
        "correlativo": correlativo or "Sin asignar",
        "aprobados": aprobados,
        "total": total,
        "porcentaje": porcentaje,
        "informes_aprobados": informes_aprobados,
        "informes_revision": informes_revision,
        "informes_pendientes": informes_pendientes,
        "informes_no_requeridos": len(no_requeridos),
        "actualizado": max(fechas) if fechas else None,
    }
    if detalle:
        iconos_item = {"aprobado": "✓", "rechazado": "×", "borrador": "○"}
        data["items"] = [
            {
                "numero": item.numero,
                "titulo": item.titulo,
                "estado": item.get_estado_item_display(),
                "estado_clave": item.estado_item,
                "icono": iconos_item[item.estado_item],
            }
            for item in items
        ]
        meses = []
        for mes in range(1, 13):
            informe = informes.get(mes)
            if mes in no_requeridos:
                clave, etiqueta, icono = "no_requerido", "No requerido", "—"
            elif informe and informe.estado == InformeMensual.ESTADO_APROBADO:
                clave, etiqueta, icono = "aprobado", "Aprobado", "✓"
            elif informe and informe.estado == InformeMensual.ESTADO_RECHAZADO:
                clave, etiqueta, icono = "rechazado", "Rechazado", "×"
            elif informe and informe.estado == InformeMensual.ESTADO_EN_REVISION:
                clave, etiqueta, icono = "revision", "En revisión", "!"
            else:
                clave, etiqueta, icono = "pendiente", "Pendiente", "○"
            meses.append({"mes": PUBLIC_MONTHS[mes - 1], "estado": etiqueta, "estado_clave": clave, "icono": icono})
        data["meses"] = meses
    return data


def _asociaciones_publicas_queryset():
    items = ItemChecklistCAIMUS.objects.filter(activo=True).only(
        "expediente_id", "numero", "titulo", "estado_item", "fecha_actualizacion"
    ).order_by("numero")
    expedientes = ExpedienteCAIMUS.objects.only(
        "id", "asociacion_id", "estado", "actualizado_en"
    ).prefetch_related(
        Prefetch("items", queryset=items, to_attr="items_publicos"),
        Prefetch(
            "resolucion",
            queryset=ResolucionExpediente.objects.only("expediente_id", "correlativo"),
            to_attr="resolucion_publica",
        ),
    )
    informes = InformeMensual.objects.only("asociacion_id", "mes", "estado", "actualizado_en")
    configs = ConfiguracionInformeAnio.objects.filter(activo=True).only("anio_id", "mes", "requerido")
    return Asociacion.objects.filter(activo=True, anio__activo=True).select_related("anio").only(
        "id", "nombre", "codigo", "anio_id", "anio__anio"
    ).prefetch_related(
        Prefetch("expediente_caimus", queryset=expedientes, to_attr="expedientes_publicos"),
        Prefetch("informes_mensuales", queryset=informes, to_attr="informes_publicos"),
        Prefetch("anio__configuracion_informes", queryset=configs, to_attr="config_publica"),
    )


def asociaciones_publicas(request):
    """Read-only transparency portal; deliberately has no authentication decorator."""
    anios = list(Anio.objects.filter(activo=True).values_list("anio", flat=True))
    anio = request.GET.get("anio", "").strip()
    busqueda = request.GET.get("q", "").strip()[:100]
    estado = request.GET.get("estado", "").strip()
    queryset = _asociaciones_publicas_queryset()
    if anio.isdigit() and int(anio) in anios:
        queryset = queryset.filter(anio__anio=int(anio))
    else:
        anio = ""
    if busqueda:
        queryset = queryset.filter(Q(nombre__icontains=busqueda) | Q(codigo__icontains=busqueda))
    asociaciones = [_datos_publicos_asociacion(obj) for obj in queryset.order_by("nombre")]
    if estado:
        asociaciones = [obj for obj in asociaciones if obj["estado_clave"] == estado]
    resumen = {
        "asociaciones": len(asociaciones),
        "aprobados": sum(obj["estado_clave"] == "aprobado" for obj in asociaciones),
        "revision": sum(obj["estado_clave"] == "revision" for obj in asociaciones),
        "pendientes": sum(obj["estado_clave"] in {"sin_iniciar", "borrador", "rechazado"} for obj in asociaciones),
        "informes": sum(obj["informes_aprobados"] for obj in asociaciones),
    }
    return render(request, "asociaciones_app/publico/lista.html", {
        "asociaciones": asociaciones, "anios": anios, "anio_seleccionado": anio,
        "busqueda": busqueda, "estado_seleccionado": estado, "resumen": resumen,
    })


def asociacion_publica_detalle(request, pk):
    asociacion = get_object_or_404(_asociaciones_publicas_queryset(), pk=pk)
    return render(request, "asociaciones_app/publico/detalle.html", {
        "asociacion": _datos_publicos_asociacion(asociacion, detalle=True),
    })


def _generar_qr_validacion_base64(validacion_url):
    qr = qrcode.QRCode(box_size=4, border=2)
    qr.add_data(validacion_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return f"data:image/png;base64,{qr_base64}"


def _contexto_institucional_constancia(request):
    institucion = Institucion.objects.first()
    logo_secundario_url = None
    if institucion and institucion.logo2:
        logo_secundario_url = request.build_absolute_uri(institucion.logo2.url)
    return {
        "institucion": institucion,
        "logo_secundario_url": logo_secundario_url,
        "footer_image_url": request.build_absolute_uri(static("assets/images/pie.png")),
    }


@asociacion_required
def dashboard(request):
    if is_admin(request.user):
        return _dashboard_admin(request)
    if is_asociacion(request.user):
        return _dashboard_asociacion(request)
    raise PermissionDenied


@asociacion_required
def asociaciones_inicio(request):
    if is_admin(request.user):
        return _dashboard_admin(request)
    if is_asociacion(request.user):
        return _dashboard_asociacion(request)
    raise PermissionDenied


def _dashboard_admin(request):
    anios_disponibles = Anio.objects.filter(activo=True).order_by("-anio")
    anio_param = request.GET.get("anio")
    anio_seleccionado = anios_disponibles.filter(anio=anio_param).first() if anio_param else None
    if anio_seleccionado is None:
        anio_seleccionado = Anio.objects.filter(activo=True).order_by("-anio").first() or anios_disponibles.first()

    asociaciones = Asociacion.objects.select_related("anio")
    expedientes = ExpedienteCAIMUS.objects.select_related("asociacion", "asociacion__anio")
    informes = InformeMensual.objects.select_related("asociacion", "asociacion__anio")
    if anio_seleccionado:
        asociaciones = asociaciones.filter(anio=anio_seleccionado)
        expedientes = expedientes.filter(asociacion__anio=anio_seleccionado)
        informes = informes.filter(asociacion__anio=anio_seleccionado)

    expedientes_por_estado = {estado: expedientes.filter(estado=estado).count() for estado, _label in ExpedienteCAIMUS.ESTADOS}
    expedientes_aprobados_validos = 0
    for expediente in expedientes.prefetch_related("items"):
        if expediente.estado == ExpedienteCAIMUS.ESTADO_APROBADO and expediente_items_100_aprobados(expediente):
            expedientes_aprobados_validos += 1
    expedientes_por_estado[ExpedienteCAIMUS.ESTADO_APROBADO] = expedientes_aprobados_validos
    informes_por_estado = {estado: informes.filter(estado=estado).count() for estado, _label in InformeMensual.ESTADOS}

    progresos = []
    for expediente in expedientes.prefetch_related("items"):
        stats = expediente.progress_stats()
        if stats["total"] > 0:
            progresos.append(stats["percent"])
    promedio_cumplimiento = int(sum(progresos) / len(progresos)) if progresos else 0

    actividad = []
    expediente_historial = ExpedienteEstadoHistorial.objects.select_related("expediente", "expediente__asociacion")
    informe_historial = InformeEstadoHistorial.objects.select_related("informe", "informe__asociacion")
    asignaciones_historial = AsociacionUsuario.objects.select_related("asociacion", "usuario")
    if anio_seleccionado:
        expediente_historial = expediente_historial.filter(expediente__asociacion__anio=anio_seleccionado)
        informe_historial = informe_historial.filter(informe__asociacion__anio=anio_seleccionado)
        asignaciones_historial = asignaciones_historial.filter(asociacion__anio=anio_seleccionado)

    for hist in expediente_historial.order_by("-cambiado_en")[:4]:
        actividad.append(
            {
                "titulo": f"Expediente: {hist.expediente.asociacion.nombre}",
                "detalle": f"{hist.estado_anterior} → {hist.estado_nuevo}",
                "fecha": hist.cambiado_en,
                "tipo": "info",
            }
        )
    for hist in informe_historial.order_by("-cambiado_en")[:4]:
        actividad.append(
            {
                "titulo": f"Informe {hist.informe.get_mes_display()} - {hist.informe.asociacion.nombre}",
                "detalle": f"{hist.estado_anterior} → {hist.estado_nuevo}",
                "fecha": hist.cambiado_en,
                "tipo": "warning",
            }
        )
    for asignacion in asignaciones_historial.order_by("-creado_en")[:4]:
        actividad.append(
            {
                "titulo": f"Asignación de usuario en {asignacion.asociacion.nombre}",
                "detalle": asignacion.usuario.get_full_name() or asignacion.usuario.username,
                "fecha": asignacion.creado_en,
                "tipo": "success",
            }
        )
    actividad_reciente = sorted(actividad, key=lambda x: x["fecha"], reverse=True)[:8]

    aprobaciones_mes = [0] * 12
    expediente_aprobaciones_qs = ExpedienteEstadoHistorial.objects.filter(estado_nuevo=ExpedienteCAIMUS.ESTADO_APROBADO)
    informe_aprobaciones_qs = InformeEstadoHistorial.objects.filter(estado_nuevo=InformeMensual.ESTADO_APROBADO)
    if anio_seleccionado:
        expediente_aprobaciones_qs = expediente_aprobaciones_qs.filter(expediente__asociacion__anio=anio_seleccionado)
        informe_aprobaciones_qs = informe_aprobaciones_qs.filter(informe__asociacion__anio=anio_seleccionado)

    for registro in (
        expediente_aprobaciones_qs
        .values("cambiado_en__month")
        .annotate(total=Count("id"))
    ):
        if registro["cambiado_en__month"]:
            aprobaciones_mes[registro["cambiado_en__month"] - 1] += registro["total"]
    for registro in (
        informe_aprobaciones_qs
        .values("cambiado_en__month")
        .annotate(total=Count("id"))
    ):
        if registro["cambiado_en__month"]:
            aprobaciones_mes[registro["cambiado_en__month"] - 1] += registro["total"]

    asociaciones_resumen = []
    resumen_informes_admin = {"requeridos": 0, "no_requeridos": 0, "pendientes": 0, "aprobados": 0}
    for asociacion in asociaciones:
        expediente = getattr(asociacion, "expediente_caimus", None)
        resumen_asoc = resumen_informes_asociacion(asociacion)
        asociaciones_resumen.append(
            {
                "asociacion": asociacion,
                "estado_expediente": expediente.estado if expediente else "SIN_EXPEDIENTE",
                "informes_aprobados": resumen_asoc["aprobados"],
                "informes_pendientes": resumen_asoc["pendientes"],
                "informes_no_requeridos": resumen_asoc["no_requeridos"],
            }
        )
        resumen_informes_admin["requeridos"] += resumen_asoc["requeridos"]
        resumen_informes_admin["no_requeridos"] += resumen_asoc["no_requeridos"]
        resumen_informes_admin["pendientes"] += resumen_asoc["pendientes"]
        resumen_informes_admin["aprobados"] += resumen_asoc["aprobados"]

    chart_payload = {
        "expedientesEstado": [
            expedientes_por_estado[ExpedienteCAIMUS.ESTADO_APROBADO],
            expedientes_por_estado[ExpedienteCAIMUS.ESTADO_EN_REVISION],
            expedientes_por_estado[ExpedienteCAIMUS.ESTADO_RECHAZADO],
            expedientes_por_estado[ExpedienteCAIMUS.ESTADO_BORRADOR],
        ],
        "informesEstado": [
            informes_por_estado[InformeMensual.ESTADO_APROBADO],
            informes_por_estado[InformeMensual.ESTADO_EN_REVISION],
            informes_por_estado[InformeMensual.ESTADO_RECHAZADO],
            informes_por_estado[InformeMensual.ESTADO_BORRADOR],
        ],
        "aprobacionesMes": aprobaciones_mes,
        "cumplimientoPromedio": promedio_cumplimiento,
    }

    resumen_inbox = resumen_dashboard_admin(anio=anio_seleccionado)

    asociaciones_resumen = asociaciones_resumen[:12]
    context = {
        "es_admin_dashboard": True,
        "kpis": {
            "total_anios": Anio.objects.filter(activo=True).count(),
            "total_asociaciones": asociaciones.count(),
            "total_usuarios_asignados": AsociacionUsuario.objects.filter(activo=True).count(),
            "total_expedientes": expedientes.count(),
            "expedientes_aprobados": expedientes_por_estado[ExpedienteCAIMUS.ESTADO_APROBADO],
            "expedientes_en_revision": expedientes_por_estado[ExpedienteCAIMUS.ESTADO_EN_REVISION],
            "expedientes_rechazados": expedientes_por_estado[ExpedienteCAIMUS.ESTADO_RECHAZADO],
            "expedientes_borrador": expedientes_por_estado[ExpedienteCAIMUS.ESTADO_BORRADOR],
            "total_informes": informes.count(),
            "informes_aprobados": informes_por_estado[InformeMensual.ESTADO_APROBADO],
            "informes_pendientes": resumen_informes_admin["pendientes"],
            "informes_requeridos": resumen_informes_admin["requeridos"],
            "informes_no_requeridos": resumen_informes_admin["no_requeridos"],
            "informes_aprobados_reales": resumen_informes_admin["aprobados"],
            "resoluciones_emitidas": ResolucionExpediente.objects.count() + ResolucionInformeMensual.objects.count(),
            "promedio_cumplimiento": promedio_cumplimiento,
        },
        "notificaciones_recientes": NotificacionAsociacion.objects.select_related("asociacion").filter(
            asociacion__anio=anio_seleccionado
        )[:8] if anio_seleccionado else NotificacionAsociacion.objects.select_related("asociacion").all()[:8],
        "alertas_nuevas_recientes": resumen_inbox["alertas_no_leidas_recientes"],
        "total_alertas_nuevas": resumen_inbox["total_alertas_no_leidas"],
        "anios_disponibles": anios_disponibles,
        "anio_seleccionado": anio_seleccionado,
        "asociaciones_resumen": asociaciones_resumen,
        "chart_payload": chart_payload,
        "meses_labels": [month_name[i] for i in range(1, 13)],
    }
    return render(request, "asociaciones_app/dashboard.html", context)


@login_required
@admin_required
def exportar_resumen_asociaciones_excel(request):
    anio_param = request.GET.get("anio")
    asociaciones = Asociacion.objects.select_related("anio")
    if anio_param:
        asociaciones = asociaciones.filter(anio__anio=anio_param)

    resumen = []
    for asociacion in asociaciones:
        expediente = getattr(asociacion, "expediente_caimus", None)
        resumen_asoc = resumen_informes_asociacion(asociacion)
        resumen.append(
            {
                "asociacion": asociacion.nombre,
                "anio": asociacion.anio.anio,
                "expediente": expediente.estado if expediente else "SIN_EXPEDIENTE",
                "informes_aprobados": resumen_asoc["aprobados"],
                "informes_pendientes": resumen_asoc["pendientes"],
                "informes_no_requeridos": resumen_asoc["no_requeridos"],
            }
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Asociaciones"
    headers = ["Asociación", "Año", "Expediente", "Informes aprobados", "Informes pendientes", "Informes no requeridos", "Fecha de descarga"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    if resumen:
        for fila in resumen:
            ws.append([
                fila["asociacion"],
                fila["anio"],
                fila["expediente"],
                fila["informes_aprobados"],
                fila["informes_pendientes"],
                fila["informes_no_requeridos"],
                timezone.now().strftime("%d/%m/%Y %H:%M"),
            ])
    else:
        ws.append(["Sin registros", "", "", "", "", "", timezone.now().strftime("%d/%m/%Y %H:%M")])

    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 24

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"resumen_asociaciones_{anio_param}.xlsx" if anio_param else "resumen_asociaciones.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _dashboard_asociacion(request):
    asociaciones_usuario = get_asociaciones_usuario(request.user).select_related("anio")
    if not asociaciones_usuario.exists():
        raise PermissionDenied

    anios_disponibles = Anio.objects.filter(asociaciones__in=asociaciones_usuario).distinct().order_by("-anio")
    anio_param = request.GET.get("anio")
    anio_seleccionado = anios_disponibles.filter(anio=anio_param).first() if anio_param else None
    if anio_seleccionado is None:
        anio_seleccionado = anios_disponibles.first()

    asociaciones = asociaciones_usuario
    if anio_seleccionado:
        asociaciones = asociaciones.filter(anio=anio_seleccionado)
    if not asociaciones.exists():
        raise PermissionDenied

    asociacion_param = request.GET.get("asociacion")
    asociacion_seleccionada = None
    if asociacion_param:
        asociacion_seleccionada = asociaciones.filter(pk=asociacion_param).first()
        if asociacion_seleccionada is None:
            raise PermissionDenied
    else:
        asociacion_seleccionada = asociaciones.first()

    expedientes = ExpedienteCAIMUS.objects.filter(asociacion=asociacion_seleccionada).select_related("asociacion")
    informes = InformeMensual.objects.filter(asociacion=asociacion_seleccionada).select_related("asociacion")
    notificaciones = NotificacionAsociacion.objects.filter(asociacion=asociacion_seleccionada).select_related("asociacion")

    total_items = 0
    items_completos = 0
    for expediente in expedientes.prefetch_related("items"):
        stats = expediente.progress_stats()
        total_items += stats["total"]
        items_completos += stats["done"]
    items_pendientes = max(total_items - items_completos, 0)
    cumplimiento = int((items_completos / total_items) * 100) if total_items else 0

    asociacion_principal = asociacion_seleccionada
    expediente_principal = expedientes.first() if asociacion_principal else None
    informes_principal = informes.order_by("mes") if asociacion_principal else InformeMensual.objects.none()
    if asociacion_principal:
        config_por_mes = {c.mes: c for c in obtener_configuracion_informes_anio(asociacion_principal.anio)["configuraciones"]}
        for informe in informes_principal:
            informe.es_requerido = config_por_mes.get(informe.mes).requerido if config_por_mes.get(informe.mes) else True

    resumen_asoc = resumen_informes_asociacion(asociacion_seleccionada)
    chart_payload = {
        "expedienteProgreso": [items_completos, items_pendientes],
        "informesResumen": [
            resumen_asoc["aprobados"],
            resumen_asoc["pendientes"],
        ],
        "cumplimiento": cumplimiento,
    }

    context = {
        "es_admin_dashboard": False,
        "mis_asociaciones": asociaciones,
        "asociaciones_disponibles": asociaciones,
        "asociacion_seleccionada": asociacion_seleccionada,
        "asociacion_principal": asociacion_principal,
        "expediente_principal": expediente_principal,
        "informes_principal": informes_principal,
        "kpis": {
            "total_mis_asociaciones": asociaciones_usuario.count(),
            "expediente_estado": expediente_principal.estado if expediente_principal else "SIN_EXPEDIENTE",
            "expediente_total_items": total_items,
            "expediente_items_completos": items_completos,
            "expediente_items_pendientes": items_pendientes,
            "informes_aprobados": resumen_asoc["aprobados"],
            "informes_pendientes": resumen_asoc["pendientes"],
            "informes_requeridos": resumen_asoc["requeridos"],
            "informes_no_requeridos": resumen_asoc["no_requeridos"],
            "alertas_no_leidas": notificaciones.filter(leida=False).count(),
            "cumplimiento": cumplimiento,
        },
        "notificaciones_recientes": notificaciones.order_by("-creada_en")[:8],
        "anios_disponibles": anios_disponibles,
        "anio_seleccionado": anio_seleccionado,
        "chart_payload": chart_payload,
        "resumen_informes": resumen_asoc,
    }
    return render(request, "asociaciones_app/dashboard.html", context)


@admin_required
def anio_list(request):
    anios = Anio.objects.all()
    return render(request, "asociaciones_app/anio_list.html", {"anios": anios})


@admin_required
def anio_informes_config(request, pk):
    anio = get_object_or_404(Anio, pk=pk)
    asegurar_configuracion_informes_anio(anio, request.user)
    configs = list(anio.configuracion_informes.order_by("mes"))
    if request.method == "POST":
        meses_requeridos = {int(v) for v in request.POST.getlist("mes_requerido")}
        for config in configs:
            config.requerido = config.mes in meses_requeridos
            config.actualizado_por = request.user
        ConfiguracionInformeAnio.objects.bulk_update(configs, ["requerido", "actualizado_por", "actualizado_en"])
        for asociacion in anio.asociaciones.all():
            crear_notificacion_asociacion(
                asociacion=asociacion,
                titulo="Configuración de informes actualizada",
                mensaje="La configuración de informes mensuales del año fue actualizada.",
                tipo=NotificacionAsociacion.TIPO_INFO,
                creada_por=request.user,
                enlace=reverse("asociaciones:informes_mensuales", args=[asociacion.pk]),
            )
        messages.success(request, "Configuración de informes actualizada correctamente.")
        return redirect("asociaciones:anio_informes_config", pk=anio.pk)
    return render(request, "asociaciones_app/anio_informes_config.html", {"anio": anio, "configs": configs})


@admin_required
def anio_create(request):
    if request.method == "POST":
        form = AnioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Año creado correctamente.")
            return redirect("asociaciones:anios_list")
    else:
        form = AnioForm()
    return render(request, "asociaciones_app/anio_form.html", {"form": form, "titulo": "Nuevo año"})


@admin_required
def anio_edit(request, pk):
    anio = get_object_or_404(Anio, pk=pk)
    if request.method == "POST":
        form = AnioForm(request.POST, instance=anio)
        if form.is_valid():
            form.save()
            messages.success(request, "Año actualizado correctamente.")
            return redirect("asociaciones:anios_list")
    else:
        form = AnioForm(instance=anio)
    return render(request, "asociaciones_app/anio_form.html", {"form": form, "titulo": "Editar año"})


@admin_required
def anio_checklist(request, pk):
    anio = get_object_or_404(Anio, pk=pk)
    formset = ChecklistAnioItemFormSet(instance=anio, prefix="checklist")
    return render(
        request,
        "asociaciones_app/anio_checklist_form.html",
        {
            "anio": anio,
            "formset": formset,
        },
    )


@admin_required
@require_POST
def anio_checklist_guardar(request, pk):
    anio = get_object_or_404(Anio, pk=pk)
    formset = ChecklistAnioItemFormSet(request.POST, instance=anio, prefix="checklist")
    if formset.is_valid():
        formset.save()
        for expediente in ExpedienteCAIMUS.objects.filter(asociacion__anio=anio).select_related("asociacion", "asociacion__anio"):
            crear_items_expediente(expediente)
            crear_notificacion_asociacion(
                asociacion=expediente.asociacion,
                titulo="Checklist anual actualizado",
                mensaje=f"Se actualizaron los requisitos del checklist del año {anio.anio}.",
                tipo=NotificacionAsociacion.TIPO_INFO,
                creada_por=request.user,
                enlace=reverse("asociaciones:expediente_caimus", args=[expediente.asociacion.pk]),
            )
        messages.success(request, "Checklist del año guardado y sincronizado correctamente.")
        return redirect("asociaciones:anio_checklist", pk=anio.pk)
    messages.error(request, "Revise los datos del checklist antes de guardar.")
    return render(
        request,
        "asociaciones_app/anio_checklist_form.html",
        {
            "anio": anio,
            "formset": formset,
        },
        status=400,
    )


@admin_required
def asociacion_list(request, anio_id):
    anio = get_object_or_404(Anio, pk=anio_id)
    asociaciones = anio.asociaciones.all()
    return render(
        request,
        "asociaciones_app/asociacion_list.html",
        {"anio": anio, "asociaciones": asociaciones},
    )


@admin_required
def asociacion_create(request, anio_id):
    anio = get_object_or_404(Anio, pk=anio_id)
    if request.method == "POST":
        form = AsociacionForm(request.POST)
        if form.is_valid():
            asociacion = form.save()
            messages.success(request, "Asociación creada correctamente.")
            return redirect("asociaciones:asociacion_list", anio_id=asociacion.anio_id)
    else:
        form = AsociacionForm(initial={"anio": anio})
    return render(
        request,
        "asociaciones_app/asociacion_form.html",
        {"form": form, "anio": anio, "titulo": "Nueva asociación"},
    )


@admin_required
def asociacion_edit(request, pk):
    asociacion = get_object_or_404(Asociacion, pk=pk)
    if request.method == "POST":
        form = AsociacionForm(request.POST, instance=asociacion)
        if form.is_valid():
            form.save()
            messages.success(request, "Asociación actualizada correctamente.")
            return redirect("asociaciones:asociacion_list", anio_id=asociacion.anio_id)
    else:
        form = AsociacionForm(instance=asociacion)
    return render(
        request,
        "asociaciones_app/asociacion_form.html",
        {"form": form, "anio": asociacion.anio, "titulo": "Editar asociación"},
    )


@admin_required
def asociacion_usuarios(request, pk):
    asociacion = get_object_or_404(Asociacion, pk=pk)
    if request.method == "POST":
        form = AsociacionUsuarioForm(request.POST, asociacion_actual=asociacion)
        if form.is_valid():
            asignacion = form.save(commit=False)
            asignacion.asociacion = asociacion
            asignacion.save()
            messages.success(request, "Usuario asignado correctamente.")
            return redirect("asociaciones:asociacion_usuarios", pk=asociacion.pk)
    else:
        form = AsociacionUsuarioForm(asociacion_actual=asociacion)
    asignaciones = asociacion.usuarios.select_related("usuario")
    return render(
        request,
        "asociaciones_app/asociacion_usuarios.html",
        {"asociacion": asociacion, "form": form, "asignaciones": asignaciones},
    )


@asociacion_required
def mis_asociaciones(request):
    if is_admin(request.user):
        asociaciones = Asociacion.objects.all()
    elif is_asociacion(request.user):
        asociaciones = get_asociaciones_usuario(request.user)
    else:
        raise PermissionDenied
    for asociacion in asociaciones:
        asociacion.notificaciones_recientes = list(asociacion.notificaciones.all()[:3])
        asociacion.notificaciones_pendientes = asociacion.notificaciones.filter(leida=False).count()
    return render(
        request,
        "asociaciones_app/mis_asociaciones.html",
        {"asociaciones": asociaciones, "es_admin": is_admin(request.user)},
    )


@asociacion_required
def expediente_caimus(request, pk):
    asociacion = get_object_or_404(Asociacion, pk=pk)
    if not user_has_asociacion_access(request.user, asociacion):
        raise PermissionDenied

    expediente, creado = ExpedienteCAIMUS.objects.get_or_create(
        asociacion=asociacion,
        defaults={"creado_por": request.user, "actualizado_por": request.user},
    )
    crear_items_expediente(expediente)
    if not asociacion.anio.checklist_items.filter(activo=True).exists():
        messages.warning(
            request,
            "El año no tenía checklist configurado. Se aplicó una plantilla base para continuar.",
        )

    if request.method == "POST":
        form = ExpedienteCAIMUSForm(request.POST, instance=expediente)
        formset = ItemChecklistFormSet(
            request.POST,
            request.FILES,
            instance=expediente,
            queryset=expediente.items.filter(activo=True).order_by("numero"),
        )
        if form.is_valid() and formset.is_valid():
            expediente = form.save(commit=False)
            expediente.actualizado_por = request.user
            expediente.save()
            formset.save()
            if is_admin(request.user):
                crear_notificacion_asociacion(
                    asociacion=asociacion,
                    titulo="Expediente actualizado",
                    mensaje="El administrador actualizó información del expediente CAIMUS.",
                    tipo=NotificacionAsociacion.TIPO_INFO,
                    creada_por=request.user,
                    enlace=reverse("asociaciones:expediente_caimus", args=[asociacion.pk]),
                )
            if request.POST.get("save_item"):
                messages.success(request, "Observación guardada correctamente.")
            else:
                messages.success(request, "Datos guardados correctamente.")
            return redirect("asociaciones:expediente_caimus", pk=asociacion.pk)
    else:
        form = ExpedienteCAIMUSForm(instance=expediente)
        formset = ItemChecklistFormSet(instance=expediente, queryset=expediente.items.filter(activo=True).order_by("numero"))

    progress = expediente.progress_stats()
    items_revision_timeline = (
        expediente.items.filter(activo=True)
        .select_related("aprobado_por", "rechazado_por")
        .prefetch_related("historial", "historial__usuario")
        .order_by("numero")
    )
    for form_item in formset.forms:
        form_item.instance.timeline_eventos_visibles = construir_timeline_eventos_visibles_item(form_item.instance)
    for item_timeline in items_revision_timeline:
        item_timeline.timeline_eventos_visibles = construir_timeline_eventos_visibles_item(item_timeline)
    expediente_completo = expediente_esta_completo(expediente)
    todos_items_aprobados = expediente_items_100_aprobados(expediente)
    estado_aprobado_valido = expediente.estado == ExpedienteCAIMUS.ESTADO_APROBADO and todos_items_aprobados
    estado_visual_expediente = expediente.estado if expediente.estado != ExpedienteCAIMUS.ESTADO_APROBADO else (
        ExpedienteCAIMUS.ESTADO_APROBADO if estado_aprobado_valido else ExpedienteCAIMUS.ESTADO_EN_REVISION
    )
    puede_descargar_resolucion = (
        is_asociacion(request.user)
        and user_has_expediente_access(request.user, expediente)
        and estado_aprobado_valido
        and expediente_completo
    )

    return render(
        request,
        "asociaciones_app/expediente_caimus_form.html",
        {
            "asociacion": asociacion,
            "expediente": expediente,
            "form": form,
            "formset": formset,
            "progress": progress,
            "total_items": progress["total"],
            "items_subidos": progress["done"],
            "items_aprobados": progress["approved"],
            "porcentaje_subidos": progress["percent"],
            "porcentaje_aprobados": progress["approved_percent"],
            "items_revision_timeline": items_revision_timeline,
            "es_admin": is_admin(request.user),
            "es_asociacion": is_asociacion(request.user),
            "expediente_completo": expediente_completo,
            "todos_items_aprobados": todos_items_aprobados,
            "estado_aprobado_valido": estado_aprobado_valido,
            "estado_visual_expediente": estado_visual_expediente,
            "puede_descargar_resolucion": puede_descargar_resolucion,
            "puede_descargar_trazabilidad": is_admin(request.user) and user_has_expediente_access(request.user, expediente),
        },
    )


@asociacion_required
@require_POST
def expediente_sync_checklist(request, pk):
    if not is_admin(request.user):
        raise PermissionDenied
    asociacion = get_object_or_404(Asociacion, pk=pk)
    expediente = get_object_or_404(ExpedienteCAIMUS, asociacion=asociacion)
    crear_items_expediente(expediente)
    messages.success(request, "Checklist del expediente sincronizado con el año.")
    return redirect("asociaciones:expediente_caimus", pk=asociacion.pk)


@asociacion_required
@require_POST
def expediente_enviar_revision(request, pk):
    asociacion = get_object_or_404(Asociacion, pk=pk)
    if not user_has_asociacion_access(request.user, asociacion):
        raise PermissionDenied
    if is_admin(request.user):
        raise PermissionDenied
    expediente = get_object_or_404(ExpedienteCAIMUS, asociacion=asociacion)
    if expediente.estado not in [ExpedienteCAIMUS.ESTADO_BORRADOR, ExpedienteCAIMUS.ESTADO_RECHAZADO]:
        messages.warning(request, "El expediente ya fue enviado a revisión.")
        return redirect("asociaciones:expediente_caimus", pk=asociacion.pk)
    if not expediente_esta_completo(expediente):
        messages.error(request, "Debes completar todos los documentos para enviar el expediente a revisión.")
        return redirect("asociaciones:expediente_caimus", pk=asociacion.pk)
    expediente.estado = ExpedienteCAIMUS.ESTADO_EN_REVISION
    expediente.actualizado_por = request.user
    expediente.save(update_fields=["estado", "actualizado_por", "actualizado_en"])
    crear_notificacion_admin(
        titulo="Expediente enviado a revisión",
        mensaje=f"La asociación {asociacion.nombre} envió su expediente a revisión.",
        tipo=NotificacionAdmin.TIPO_WARNING,
        creada_por=request.user,
        enlace=reverse("asociaciones:expediente_caimus", args=[asociacion.pk]),
        asociacion=asociacion,
    )
    crear_entrada_revision_admin(
        tipo=EntradaRevisionAdmin.TIPO_EXPEDIENTE,
        titulo="Expediente enviado a revisión",
        mensaje=f"La asociación {asociacion.nombre} envió su expediente a revisión.",
        creada_por=request.user,
        enlace=reverse("asociaciones:expediente_caimus", args=[asociacion.pk]),
        asociacion=asociacion,
        expediente=expediente,
    )
    messages.success(request, "Expediente enviado a revisión correctamente.")
    return redirect("asociaciones:expediente_caimus", pk=asociacion.pk)


@asociacion_required
@require_POST
def item_upload(request, expediente_id, item_id):
    expediente = get_object_or_404(ExpedienteCAIMUS, pk=expediente_id)
    if not user_has_expediente_access(request.user, expediente):
        raise PermissionDenied
    item = get_object_or_404(expediente.items, pk=item_id)
    archivo = request.FILES.get("pdf")
    if not archivo:
        messages.error(request, "Debe seleccionar un archivo válido (PDF, XLS o XLSX).")
        return redirect("asociaciones:expediente_caimus", pk=expediente.asociacion.pk)

    extension = Path(archivo.name).suffix.lower()
    if extension not in ALLOWED_EXPEDIENTE_ITEM_EXTENSIONS:
        messages.error(request, "Formato no permitido. Formatos permitidos: PDF, XLS y XLSX.")
        return redirect("asociaciones:expediente_caimus", pk=expediente.asociacion.pk)

    if archivo.content_type and archivo.content_type not in ALLOWED_EXPEDIENTE_ITEM_CONTENT_TYPES:
        messages.error(request, "Tipo de archivo no válido. Formatos permitidos: PDF, XLS y XLSX.")
        return redirect("asociaciones:expediente_caimus", pk=expediente.asociacion.pk)

    tenia_archivo = bool(item.pdf)
    item.pdf = archivo
    item.subido_por = request.user
    try:
        item.full_clean()
    except ValidationError as exc:
        messages.error(request, "; ".join(exc.messages))
        return redirect("asociaciones:expediente_caimus", pk=expediente.asociacion.pk)
    item.save()
    HistorialItemExpediente.objects.create(
        item=item,
        usuario=request.user,
        accion=(
            HistorialItemExpediente.ACCION_ARCHIVO_ACTUALIZADO
            if tenia_archivo
            else HistorialItemExpediente.ACCION_ARCHIVO_SUBIDO
        ),
        descripcion=(
            "El usuario actualizó o reemplazó el archivo del documento."
            if tenia_archivo
            else "El usuario cargó el archivo del documento."
        ),
        archivo=item.pdf,
    )
    messages.success(request, "Archivo subido correctamente.")
    return redirect("asociaciones:expediente_caimus", pk=expediente.asociacion.pk)


@asociacion_required
@require_POST
def item_observacion(request, expediente_id, item_id):
    expediente = get_object_or_404(ExpedienteCAIMUS, pk=expediente_id)
    if not user_has_expediente_access(request.user, expediente):
        raise PermissionDenied
    item = get_object_or_404(expediente.items, pk=item_id)
    item.observaciones = request.POST.get("observaciones", "")
    item.save()
    messages.success(request, "Observación guardada correctamente.")
    return redirect("asociaciones:expediente_caimus", pk=expediente.asociacion.pk)


def _nombre_usuario(user):
    if not user:
        return ""
    return user.get_full_name() or user.username


def formatear_duracion(delta):
    """Devuelve una duración útil para lectura humana, sin exponer segundos."""
    if delta is None:
        return "No disponible"
    total_segundos = max(0, int(delta.total_seconds()))
    dias, resto = divmod(total_segundos, 86400)
    horas, resto = divmod(resto, 3600)
    minutos, _ = divmod(resto, 60)
    partes = []
    if dias:
        partes.append(f"{dias} día{'s' if dias != 1 else ''}")
    if horas:
        partes.append(f"{horas} hora{'s' if horas != 1 else ''}")
    if minutos or not partes:
        partes.append(f"{minutos} minuto{'s' if minutos != 1 else ''}")
    return ", ".join(partes)


def obtener_nombre_usuario(usuario, valor_por_defecto="No disponible"):
    """Obtiene un nombre presentable sin acceder a atributos de usuarios nulos."""
    if not usuario:
        return valor_por_defecto
    nombre_completo = usuario.get_full_name().strip()
    return nombre_completo or usuario.username or valor_por_defecto


def construir_timeline_eventos_visibles_item(item):
    historial_ordenado = list(item.historial.select_related("usuario").all().order_by("creado_en", "id"))
    evento_inicial = next((e for e in historial_ordenado if e.accion == HistorialItemExpediente.ACCION_ARCHIVO_SUBIDO), None)
    primer_evento = historial_ordenado[0] if historial_ordenado else None
    archivo_item = getattr(item, "archivo", None) or getattr(item, "pdf", None)

    if not archivo_item and not historial_ordenado:
        return []

    usuario_inicio = (
        (evento_inicial.usuario if evento_inicial and evento_inicial.usuario else None)
        or getattr(item, "subido_por", None)
        or (primer_evento.usuario if primer_evento and primer_evento.usuario else None)
    )
    fecha_inicio = (
        (evento_inicial.creado_en if evento_inicial else None)
        or getattr(item, "fecha_carga", None)
        or (primer_evento.creado_en if primer_evento else None)
        or getattr(item, "fecha_actualizacion", None)
    )
    archivo_inicio = (
        (evento_inicial.archivo if evento_inicial and evento_inicial.archivo else None)
        or getattr(item, "archivo", None)
        or getattr(item, "pdf", None)
    )

    tiene_rechazo = any(evento.accion == HistorialItemExpediente.ACCION_RECHAZADO for evento in historial_ordenado)
    observaciones_admin = [e for e in historial_ordenado if e.accion == HistorialItemExpediente.ACCION_OBSERVACION_ADMIN]
    observacion_rechazo = observaciones_admin[-1].descripcion if observaciones_admin else ""

    eventos_visibles = []
    if evento_inicial:
        eventos_visibles.append({
            "id": f"inicio-{item.id}",
            "evento_id": evento_inicial.id,
            "accion": "INICIO",
            "timeline_accion": "INICIO",
            "timeline_titulo": "Inicio",
            "timeline_subtitulo": "Carga inicial",
            "timeline_descripcion": evento_inicial.descripcion or "Inicio del proceso con la primera carga del documento.",
            "creado_en": evento_inicial.creado_en,
            "fecha": evento_inicial.creado_en,
            "usuario": evento_inicial.usuario,
            "archivo": evento_inicial.archivo,
            "es_fallback": False,
        })
    elif archivo_item:
        eventos_visibles.append({
            "id": f"inicio-fallback-{item.id}",
            "evento_id": f"fallback-{item.id}",
            "accion": "INICIO",
            "timeline_accion": "INICIO",
            "timeline_titulo": "Inicio",
            "timeline_subtitulo": "Carga inicial",
            "timeline_descripcion": "Inicio del proceso con archivo existente.",
            "creado_en": fecha_inicio,
            "fecha": fecha_inicio,
            "usuario": usuario_inicio,
            "archivo": archivo_inicio,
            "es_fallback": True,
        })

    for evento in historial_ordenado:
        if evento_inicial and evento.id == evento_inicial.id:
            continue
        if tiene_rechazo and evento.accion == HistorialItemExpediente.ACCION_OBSERVACION_ADMIN:
            continue
        if evento.accion == HistorialItemExpediente.ACCION_RECHAZADO:
            evento.observacion_admin_modal = observacion_rechazo
        eventos_visibles.append(evento)

    return eventos_visibles


@asociacion_required
@require_POST
def item_revision_estado(request, expediente_id, item_id):
    if not is_admin(request.user):
        raise PermissionDenied
    expediente = get_object_or_404(ExpedienteCAIMUS, pk=expediente_id)
    if not user_has_expediente_access(request.user, expediente):
        raise PermissionDenied
    item = get_object_or_404(expediente.items, pk=item_id, activo=True)
    accion = request.POST.get("accion")
    observacion_revision = (request.POST.get("observacion_revision") or "").strip()
    ahora = timezone.now()
    observacion_cambio = observacion_revision != (item.observacion_revision or "").strip()
    if accion == "aprobar":
        item.estado_item = item.ESTADO_APROBADO
        item.aprobado_por = request.user
        item.fecha_aprobacion = ahora
        item.rechazado_por = None
        item.fecha_rechazo = None
        item.observacion_revision = observacion_revision
        item.save(update_fields=["estado_item", "aprobado_por", "fecha_aprobacion", "rechazado_por", "fecha_rechazo", "observacion_revision"])
        if observacion_revision and observacion_cambio:
            HistorialItemExpediente.objects.create(
                item=item,
                usuario=request.user,
                accion=HistorialItemExpediente.ACCION_OBSERVACION_ADMIN,
                descripcion=observacion_revision,
            )
        HistorialItemExpediente.objects.create(
            item=item,
            usuario=request.user,
            accion=HistorialItemExpediente.ACCION_APROBADO,
            descripcion="El documento fue aprobado.",
        )
        crear_notificacion_asociacion(
            asociacion=expediente.asociacion,
            titulo="Documento aprobado",
            mensaje="Uno de los documentos de su expediente ha sido aprobado.",
            tipo=NotificacionAsociacion.TIPO_SUCCESS,
            creada_por=request.user,
            enlace=reverse("asociaciones:expediente_caimus", args=[expediente.asociacion.pk]),
        )
    elif accion == "rechazar":
        item.estado_item = item.ESTADO_RECHAZADO
        item.rechazado_por = request.user
        item.fecha_rechazo = ahora
        item.aprobado_por = None
        item.fecha_aprobacion = None
        item.observacion_revision = observacion_revision
        item.save(update_fields=["estado_item", "rechazado_por", "fecha_rechazo", "aprobado_por", "fecha_aprobacion", "observacion_revision"])
        if observacion_revision and observacion_cambio:
            HistorialItemExpediente.objects.create(
                item=item,
                usuario=request.user,
                accion=HistorialItemExpediente.ACCION_OBSERVACION_ADMIN,
                descripcion=observacion_revision,
            )
        HistorialItemExpediente.objects.create(
            item=item,
            usuario=request.user,
            accion=HistorialItemExpediente.ACCION_RECHAZADO,
            descripcion=observacion_revision or "El documento fue rechazado.",
        )
        crear_notificacion_asociacion(
            asociacion=expediente.asociacion,
            titulo="Documento rechazado",
            mensaje="Uno de los documentos de su expediente ha sido rechazado. Revise las observaciones.",
            tipo=NotificacionAsociacion.TIPO_ERROR,
            creada_por=request.user,
            enlace=reverse("asociaciones:expediente_caimus", args=[expediente.asociacion.pk]),
        )
    else:
        return JsonResponse({"ok": False, "message": "Acción inválida."}, status=400)
    return JsonResponse({"ok": True})


@asociacion_required
def informes_mensuales(request, pk):
    asociacion = get_object_or_404(Asociacion, pk=pk)
    if not user_has_asociacion_access(request.user, asociacion):
        raise PermissionDenied
    crear_informes_mensuales(asociacion, request.user)
    asegurar_configuracion_informes_anio(asociacion.anio)
    informes = asociacion.informes_mensuales.all()
    config_map = {c.mes: c for c in asociacion.anio.configuracion_informes.filter(activo=True)}
    for informe in informes:
        informe.es_requerido = config_map.get(informe.mes).requerido if config_map.get(informe.mes) else True
    resumen_informes = resumen_informes_asociacion(asociacion)
    puede_subir = is_admin(request.user) or user_has_asociacion_access(request.user, asociacion)
    return render(
        request,
        "asociaciones_app/informes_mensuales.html",
        {
            "asociacion": asociacion,
            "informes": informes,
            "es_admin": is_admin(request.user),
            "es_asociacion": is_asociacion(request.user),
            "puede_subir": puede_subir,
            "config_map": config_map,
            "resumen_informes": resumen_informes,
        },
    )


@asociacion_required
@require_POST
def informe_upload_narrativo(request, asociacion_id, mes):
    return _informe_upload_por_tipo(request, asociacion_id, mes, "narrativo")


@asociacion_required
@require_POST
def informe_upload_presupuestario(request, asociacion_id, mes):
    return _informe_upload_por_tipo(request, asociacion_id, mes, "presupuestario")


@asociacion_required
@require_POST
def informe_upload_presupuestario_excel(request, asociacion_id, mes):
    return _informe_upload_por_tipo(request, asociacion_id, mes, "presupuestario_excel")


@asociacion_required
@require_POST
def informe_upload(request, asociacion_id, mes):
    # Compatibilidad con endpoint previo: el archivo legado se guarda como narrativo.
    return _informe_upload_por_tipo(request, asociacion_id, mes, "narrativo")


def _informe_upload_por_tipo(request, asociacion_id, mes, tipo_archivo):
    asociacion = get_object_or_404(Asociacion, pk=asociacion_id)
    if not user_has_asociacion_access(request.user, asociacion):
        raise PermissionDenied
    if mes not in range(1, 13):
        messages.error(request, "Mes inválido.")
        return redirect("asociaciones:informes_mensuales", pk=asociacion.pk)
    if not informe_mes_requerido(asociacion, mes):
        messages.warning(request, "Este informe mensual no está requerido para el año seleccionado.")
        return redirect("asociaciones:informes_mensuales", pk=asociacion.pk)
    informe, _creado = InformeMensual.objects.get_or_create(
        asociacion=asociacion,
        mes=mes,
        defaults={"creado_por": request.user, "actualizado_por": request.user},
    )
    if tipo_archivo not in ["narrativo", "presupuestario", "presupuestario_excel"]:
        messages.error(request, "Tipo de archivo inválido.")
        return redirect("asociaciones:informes_mensuales", pk=asociacion.pk)
    archivo = request.FILES.get("pdf")
    if not archivo:
        messages.error(request, "Debe seleccionar un archivo PDF.")
        return redirect("asociaciones:informes_mensuales", pk=asociacion.pk)
    if tipo_archivo == "presupuestario_excel":
        tipos_excel = [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/octet-stream",
        ]
        if not archivo.name.lower().endswith(".xlsx") or archivo.content_type not in tipos_excel:
            messages.error(request, "El archivo debe ser formato Excel (.xlsx)")
            return redirect("asociaciones:informes_mensuales", pk=asociacion.pk)
    else:
        if archivo.content_type != "application/pdf":
            messages.error(request, "El archivo debe ser un PDF válido.")
            return redirect("asociaciones:informes_mensuales", pk=asociacion.pk)
    if tipo_archivo == "presupuestario_excel" and informe.estado == InformeMensual.ESTADO_APROBADO:
        messages.error(request, "No se puede subir archivo en un informe aprobado.")
        return redirect("asociaciones:informes_mensuales", pk=asociacion.pk)

    if tipo_archivo == "narrativo":
        informe.archivo_narrativo = archivo
    elif tipo_archivo == "presupuestario":
        informe.archivo_presupuestario = archivo
    else:
        informe.archivo_presupuestario_excel = archivo
    if is_asociacion(request.user):
        if informe.estado in [InformeMensual.ESTADO_RECHAZADO, InformeMensual.ESTADO_APROBADO]:
            informe.estado = InformeMensual.ESTADO_BORRADOR
    informe.observacion_admin = ""
    informe.aprobado_por = None
    informe.aprobado_en = None
    informe.actualizado_por = request.user
    try:
        informe.full_clean()
    except ValidationError as exc:
        messages.error(request, "; ".join(exc.messages))
        return redirect("asociaciones:informes_mensuales", pk=asociacion.pk)
    informe.save()
    messages.success(request, "Archivo cargado correctamente.")
    return redirect("asociaciones:informes_mensuales", pk=asociacion.pk)


@asociacion_required
@require_POST
def informe_enviar_revision(request, asociacion_id, mes):
    asociacion = get_object_or_404(Asociacion, pk=asociacion_id)
    if not user_has_asociacion_access(request.user, asociacion):
        raise PermissionDenied
    if is_admin(request.user):
        raise PermissionDenied
    informe = get_object_or_404(asociacion.informes_mensuales, mes=mes)
    if not informe_mes_requerido(asociacion, mes):
        messages.warning(request, "Este informe mensual no está requerido para el año seleccionado.")
        return redirect("asociaciones:informes_mensuales", pk=asociacion.pk)
    if informe.estado == InformeMensual.ESTADO_EN_REVISION:
        messages.warning(request, "El informe ya está en revisión.")
        return redirect("asociaciones:informes_mensuales", pk=asociacion.pk)
    if informe.estado == InformeMensual.ESTADO_APROBADO:
        messages.warning(request, "El informe ya está aprobado.")
        return redirect("asociaciones:informes_mensuales", pk=asociacion.pk)
    if not informe.tiene_archivos_completos():
        messages.error(request, "Debes cargar ambos archivos antes de enviar a revisión.")
        return redirect("asociaciones:informes_mensuales", pk=asociacion.pk)
    informe.estado = InformeMensual.ESTADO_EN_REVISION
    informe.actualizado_por = request.user
    informe.save(update_fields=["estado", "actualizado_por", "actualizado_en"])
    crear_notificacion_admin(
        titulo="Informe enviado a revisión",
        mensaje=f"La asociación {asociacion.nombre} envió a revisión el informe mensual de {informe.get_mes_display()}.",
        tipo=NotificacionAdmin.TIPO_WARNING,
        creada_por=request.user,
        enlace=f"{reverse('asociaciones:informes_mensuales', args=[asociacion.pk])}#informe-mes-{informe.mes}",
        asociacion=asociacion,
        informe=informe,
    )
    crear_entrada_revision_admin(
        tipo=EntradaRevisionAdmin.TIPO_INFORME,
        titulo="Informe enviado a revisión",
        mensaje=f"La asociación {asociacion.nombre} envió a revisión el informe mensual de {informe.get_mes_display()}.",
        creada_por=request.user,
        enlace=f"{reverse('asociaciones:informes_mensuales', args=[asociacion.pk])}#informe-mes-{informe.mes}",
        asociacion=asociacion,
        informe=informe,
    )
    messages.success(request, "Informe enviado a revisión correctamente.")
    return redirect("asociaciones:informes_mensuales", pk=asociacion.pk)


@asociacion_required
@require_POST
def informe_observacion(request, asociacion_id, mes):
    asociacion = get_object_or_404(Asociacion, pk=asociacion_id)
    if not user_has_asociacion_access(request.user, asociacion):
        raise PermissionDenied
    if is_admin(request.user):
        raise PermissionDenied
    if mes not in range(1, 13):
        messages.error(request, "Mes inválido.")
        return redirect("asociaciones:informes_mensuales", pk=asociacion.pk)
    if not informe_mes_requerido(asociacion, mes):
        messages.warning(request, "Este informe mensual no está requerido para el año seleccionado.")
        return redirect("asociaciones:informes_mensuales", pk=asociacion.pk)
    informe = get_object_or_404(asociacion.informes_mensuales, mes=mes)
    informe.observaciones_usuario = request.POST.get("observaciones", "")
    informe.actualizado_por = request.user
    informe.save()
    messages.success(request, "Observaciones guardadas correctamente.")
    return redirect("asociaciones:informes_mensuales", pk=asociacion.pk)


@login_required
@admin_required
def alerta_admin_revisar(request, pk):
    alerta = get_object_or_404(NotificacionAdmin, pk=pk)
    if not alerta.leida:
        alerta.leida = True
        alerta.save(update_fields=["leida"])
    if alerta.enlace:
        return redirect(alerta.enlace)
    return redirect("asociaciones:inicio")


@login_required
@admin_required
@require_POST
def informe_estado(request, asociacion_id, mes):
    asociacion = get_object_or_404(Asociacion, pk=asociacion_id)
    if mes not in range(1, 13):
        messages.error(request, "Mes inválido.")
        return redirect("asociaciones:informes_mensuales", pk=asociacion.pk)
    if not informe_mes_requerido(asociacion, mes):
        messages.warning(request, "Este informe mensual no está requerido para el año seleccionado.")
        return redirect("asociaciones:informes_mensuales", pk=asociacion.pk)
    informe = get_object_or_404(asociacion.informes_mensuales, mes=mes)
    estado_anterior = informe.estado
    observacion_anterior = (informe.observacion_admin or "").strip()
    estado_nuevo = request.POST.get("estado")
    if estado_nuevo not in dict(InformeMensual.ESTADOS):
        messages.error(request, "Estado inválido.")
        return redirect("asociaciones:informes_mensuales", pk=asociacion.pk)
    observacion_admin = request.POST.get("observacion_admin", "").strip()
    if estado_nuevo == InformeMensual.ESTADO_RECHAZADO and not observacion_admin:
        messages.error(request, "Debe indicar la observación del rechazo.")
        return redirect("asociaciones:informes_mensuales", pk=asociacion.pk)
    if estado_nuevo == InformeMensual.ESTADO_APROBADO and not informe.tiene_archivos_completos():
        messages.error(
            request,
            "Para aprobar el informe mensual deben cargarse tanto el informe narrativo como el presupuestario.",
        )
        return redirect("asociaciones:informes_mensuales", pk=asociacion.pk)

    informe.estado = estado_nuevo
    if estado_nuevo == InformeMensual.ESTADO_APROBADO:
        informe.aprobado_por = request.user
        informe.aprobado_en = timezone.now()
    else:
        informe.aprobado_por = None
        informe.aprobado_en = None
    informe.observacion_admin = observacion_admin
    informe.actualizado_por = request.user
    informe.save()

    InformeEstadoHistorial.objects.create(
        informe=informe,
        estado_anterior=estado_anterior,
        estado_nuevo=estado_nuevo,
        observacion=observacion_admin,
        cambiado_por=request.user,
    )
    EntradaRevisionAdmin.objects.filter(
        informe=informe,
        estado=EntradaRevisionAdmin.ESTADO_PENDIENTE,
    ).update(
        estado=EntradaRevisionAdmin.ESTADO_ATENDIDA,
        atendida_en=timezone.now(),
    )
    if estado_nuevo == InformeMensual.ESTADO_APROBADO and not hasattr(informe, "resolucion"):
        correlativo = generar_correlativo_informe(asociacion.anio.anio, informe.mes)
        ResolucionInformeMensual.objects.create(
            informe=informe,
            correlativo=correlativo,
            fecha_emision=timezone.now().date(),
            generado_por=request.user,
            contenido_snapshot={
                "asociacion": asociacion.nombre,
                "anio": asociacion.anio.anio,
                "mes": informe.get_mes_display(),
                "estado": informe.estado,
            },
        )
    if estado_nuevo == InformeMensual.ESTADO_APROBADO:
        crear_notificacion_asociacion(
            asociacion=asociacion,
            titulo="Informe mensual aprobado",
            mensaje=f"El informe mensual de {informe.get_mes_display()} fue aprobado.",
            tipo=NotificacionAsociacion.TIPO_SUCCESS,
            creada_por=request.user,
            enlace=reverse("asociaciones:informes_mensuales", args=[asociacion.pk]),
        )
    elif estado_nuevo == InformeMensual.ESTADO_RECHAZADO:
        crear_notificacion_asociacion(
            asociacion=asociacion,
            titulo="Informe mensual rechazado",
            mensaje=f"El informe mensual de {informe.get_mes_display()} fue rechazado. Revise observaciones.",
            tipo=NotificacionAsociacion.TIPO_ERROR,
            creada_por=request.user,
            enlace=reverse("asociaciones:informes_mensuales", args=[asociacion.pk]),
        )
    if observacion_anterior != observacion_admin and observacion_admin:
        crear_notificacion_asociacion(
            asociacion=asociacion,
            titulo="Nueva observación en informe mensual",
            mensaje=(
                f"El administrador agregó o actualizó una observación en el informe mensual de "
                f"{informe.get_mes_display()}. Revísala."
            ),
            tipo=NotificacionAsociacion.TIPO_WARNING,
            creada_por=request.user,
            enlace=reverse("asociaciones:informes_mensuales", args=[asociacion.pk]),
        )
    messages.success(request, "Estado actualizado correctamente.")
    return redirect("asociaciones:informes_mensuales", pk=asociacion.pk)


@asociacion_required
def informe_resolucion_pdf(request, asociacion_id, mes):
    asociacion = get_object_or_404(Asociacion, pk=asociacion_id)
    if not user_has_asociacion_access(request.user, asociacion):
        raise PermissionDenied
    if mes not in range(1, 13):
        raise PermissionDenied
    informe = get_object_or_404(asociacion.informes_mensuales, mes=mes)
    if informe.estado != InformeMensual.ESTADO_APROBADO:
        messages.error(request, "La constancia estará disponible cuando el informe sea aprobado.")
        raise PermissionDenied

    resolucion = getattr(informe, "resolucion", None)
    if resolucion is None:
        correlativo = generar_correlativo_informe(asociacion.anio.anio, informe.mes)
        resolucion = ResolucionInformeMensual.objects.create(
            informe=informe,
            correlativo=correlativo,
            fecha_emision=timezone.now().date(),
            generado_por=request.user,
            contenido_snapshot={
                "asociacion": asociacion.nombre,
                "anio": asociacion.anio.anio,
                "mes": informe.get_mes_display(),
                "estado": informe.estado,
            },
        )

    nombre_archivo = (
        f"constancia_informe_{informe.mes:02d}_{asociacion.codigo or asociacion.pk}_{asociacion.anio.anio}.pdf"
    )
    if resolucion.archivo_pdf and resolucion.archivo_pdf.storage.exists(resolucion.archivo_pdf.name):
        archivo = resolucion.archivo_pdf.open("rb")
        response = HttpResponse(archivo.read(), content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="{nombre_archivo}"'
        return response

    html = render_to_string(
        "asociaciones_app/informe_resolucion_pdf.html",
        {
            "asociacion": asociacion,
            "informe": informe,
            "resolucion": resolucion,
        },
        request=request,
    )
    pdf = HTML(string=html, base_url=request.build_absolute_uri("/")).write_pdf()
    resolucion.archivo_pdf.save(nombre_archivo, ContentFile(pdf), save=True)
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{nombre_archivo}"'
    return response


@login_required
@admin_required
def expediente_revision(request, pk):
    expediente = get_object_or_404(ExpedienteCAIMUS, pk=pk)
    estado_anterior = expediente.estado
    observacion_anterior = (expediente.observacion_admin or "").strip()

    if request.method == "POST":
        form = RevisionExpedienteForm(request.POST, instance=expediente)
        if form.is_valid():
            expediente = form.save(commit=False)
            observacion_nueva = (expediente.observacion_admin or "").strip()
            expediente.observacion_admin = observacion_nueva
            todos_items_aprobados = not expediente.items.filter(activo=True).exclude(
                estado_item=ItemChecklistCAIMUS.ESTADO_APROBADO
            ).exists()
            if expediente.estado == ExpedienteCAIMUS.ESTADO_APROBADO and (
                not todos_items_aprobados or not expediente_esta_completo(expediente)
            ):
                messages.error(
                    request,
                    "No es posible aprobar el expediente. Aún existen ítems pendientes de aprobación.",
                )
                return redirect("asociaciones:expediente_caimus", pk=expediente.asociacion.pk)
            if expediente.estado == ExpedienteCAIMUS.ESTADO_APROBADO:
                expediente.aprobado_por = request.user
                expediente.aprobado_en = timezone.now()
            else:
                expediente.aprobado_por = None
                expediente.aprobado_en = None
            expediente.actualizado_por = request.user
            expediente.save()
            EntradaRevisionAdmin.objects.filter(
                expediente=expediente,
                estado=EntradaRevisionAdmin.ESTADO_PENDIENTE,
            ).update(
                estado=EntradaRevisionAdmin.ESTADO_ATENDIDA,
                atendida_en=timezone.now(),
            )

            ExpedienteEstadoHistorial.objects.create(
                expediente=expediente,
                estado_anterior=estado_anterior,
                estado_nuevo=expediente.estado,
                observacion=expediente.observacion_admin,
                cambiado_por=request.user,
            )

            if expediente.estado == ExpedienteCAIMUS.ESTADO_APROBADO and not hasattr(expediente, "resolucion"):
                correlativo = generar_correlativo(expediente.asociacion.anio.anio)
                ResolucionExpediente.objects.create(
                    expediente=expediente,
                    correlativo=correlativo,
                    fecha_emision=timezone.now().date(),
                    generado_por=request.user,
                    contenido_snapshot={
                        "asociacion": expediente.asociacion.nombre,
                        "anio": expediente.asociacion.anio.anio,
                        "estado": expediente.estado,
                    },
                )
            if expediente.estado == ExpedienteCAIMUS.ESTADO_APROBADO:
                crear_notificacion_asociacion(
                    asociacion=expediente.asociacion,
                    titulo="Expediente aprobado",
                    mensaje="El expediente fue aprobado.",
                    tipo=NotificacionAsociacion.TIPO_SUCCESS,
                    creada_por=request.user,
                    enlace=reverse("asociaciones:expediente_caimus", args=[expediente.asociacion.pk]),
                )
            elif expediente.estado == ExpedienteCAIMUS.ESTADO_RECHAZADO:
                crear_notificacion_asociacion(
                    asociacion=expediente.asociacion,
                    titulo="Expediente rechazado",
                    mensaje="El expediente fue rechazado. Revise observaciones.",
                    tipo=NotificacionAsociacion.TIPO_ERROR,
                    creada_por=request.user,
                    enlace=reverse("asociaciones:expediente_caimus", args=[expediente.asociacion.pk]),
                )
            elif expediente.estado == ExpedienteCAIMUS.ESTADO_EN_REVISION:
                crear_notificacion_asociacion(
                    asociacion=expediente.asociacion,
                    titulo="Expediente en revisión",
                    mensaje="El administrador cambió el estado del expediente a en revisión.",
                    tipo=NotificacionAsociacion.TIPO_WARNING,
                    creada_por=request.user,
                    enlace=reverse("asociaciones:expediente_caimus", args=[expediente.asociacion.pk]),
                )
            if observacion_anterior != observacion_nueva and observacion_nueva:
                crear_notificacion_asociacion(
                    asociacion=expediente.asociacion,
                    titulo="Nueva observación en expediente",
                    mensaje="El administrador agregó o actualizó una observación en tu expediente. Revísala.",
                    tipo=NotificacionAsociacion.TIPO_WARNING,
                    creada_por=request.user,
                    enlace=reverse("asociaciones:expediente_caimus", args=[expediente.asociacion.pk]),
                )

            messages.success(request, "Estado actualizado correctamente.")
            return redirect("asociaciones:expediente_caimus", pk=expediente.asociacion.pk)
    else:
        form = RevisionExpedienteForm(instance=expediente)

    return render(
        request,
        "asociaciones_app/expediente_revision.html",
        {"expediente": expediente, "form": form},
    )


@login_required
@admin_required
def bandeja_revision(request):
    tipo = request.GET.get("tipo")
    estado = request.GET.get("estado")
    anio_id = request.GET.get("anio")
    tipo_filtro = tipo if tipo in [EntradaRevisionAdmin.TIPO_EXPEDIENTE, EntradaRevisionAdmin.TIPO_INFORME] else None
    estado_filtro = estado if estado in [EntradaRevisionAdmin.ESTADO_PENDIENTE, EntradaRevisionAdmin.ESTADO_ATENDIDA] else None
    entradas = obtener_entradas_bandeja_admin(anio=anio_id, estado=estado_filtro, tipo=tipo_filtro)
    anios = Anio.objects.order_by("-anio")
    return render(
        request,
        "asociaciones_app/bandeja_revision.html",
        {
            "entradas": entradas,
            "anios": anios,
            "tipo": tipo,
            "estado": estado,
            "anio_id": anio_id,
            "tipos": EntradaRevisionAdmin.TIPOS,
            "estados": EntradaRevisionAdmin.ESTADOS,
            "pendientes_total": entradas.filter(estado=EntradaRevisionAdmin.ESTADO_PENDIENTE).count(),
        },
    )


@login_required
@admin_required
@require_POST
def bandeja_marcar_atendida(request, pk):
    entrada = get_object_or_404(EntradaRevisionAdmin, pk=pk)
    if entrada.estado != EntradaRevisionAdmin.ESTADO_ATENDIDA:
        entrada.estado = EntradaRevisionAdmin.ESTADO_ATENDIDA
        entrada.atendida_en = timezone.now()
        entrada.save(update_fields=["estado", "atendida_en"])
        messages.success(request, "La entrada fue marcada como atendida.")
    return redirect("asociaciones:bandeja_revision")


@admin_required
def asignaciones_list(request):
    anio_id = request.GET.get("anio")
    asignaciones = AsociacionUsuario.objects.select_related("asociacion", "asociacion__anio", "usuario")
    if anio_id:
        asignaciones = asignaciones.filter(asociacion__anio_id=anio_id)
    anios = Anio.objects.all()
    return render(
        request,
        "asociaciones_app/asignaciones_list.html",
        {"asignaciones": asignaciones, "anios": anios, "anio_id": anio_id},
    )


def _datos_item_trazabilidad(item, fecha_generacion):
    eventos = list(item.historial.all())
    cargas = [e for e in eventos if e.accion in {
        HistorialItemExpediente.ACCION_ARCHIVO_SUBIDO,
        HistorialItemExpediente.ACCION_ARCHIVO_ACTUALIZADO,
    }]
    primera_carga_evento = next(
        (e for e in cargas if e.accion == HistorialItemExpediente.ACCION_ARCHIVO_SUBIDO),
        cargas[0] if cargas else None,
    )
    primera_carga = primera_carga_evento.creado_en if primera_carga_evento else (item.fecha_carga if item.pdf else None)
    primera_revision = next((e for e in eventos if e.accion in {
        HistorialItemExpediente.ACCION_EN_REVISION,
        HistorialItemExpediente.ACCION_OBSERVACION_ADMIN,
        HistorialItemExpediente.ACCION_APROBADO,
        HistorialItemExpediente.ACCION_RECHAZADO,
    }), None)
    decisiones = [e for e in eventos if e.accion in {
        HistorialItemExpediente.ACCION_APROBADO,
        HistorialItemExpediente.ACCION_RECHAZADO,
    }]
    decision = decisiones[-1] if decisiones else None
    aprobacion = next((e for e in reversed(eventos) if e.accion == HistorialItemExpediente.ACCION_APROBADO), None)
    observaciones = [e for e in eventos if e.accion == HistorialItemExpediente.ACCION_OBSERVACION_ADMIN]
    actualizaciones = [e for e in eventos if e.accion == HistorialItemExpediente.ACCION_ARCHIVO_ACTUALIZADO]
    ciclos = sum(
        1 for actualizacion in actualizaciones
        if any(e.creado_en < actualizacion.creado_en and e.accion in {
            HistorialItemExpediente.ACCION_OBSERVACION_ADMIN,
            HistorialItemExpediente.ACCION_RECHAZADO,
        } for e in eventos)
    )
    respuestas = []
    for observacion in observaciones:
        siguiente = next((e for e in actualizaciones if e.creado_en > observacion.creado_en), None)
        respuestas.append({
            "observacion": observacion,
            "actualizacion": siguiente,
            "duracion": formatear_duracion(siguiente.creado_en - observacion.creado_en) if siguiente else "Pendiente de respuesta",
        })
    fin = aprobacion.creado_en if aprobacion else fecha_generacion
    ultima = eventos[-1] if eventos else None
    for evento in eventos:
        evento.nombre_archivo = Path(evento.archivo.name).name if evento.archivo else "—"
        evento.tipo_archivo = Path(evento.nombre_archivo).suffix.lstrip(".").upper() or "—"
        evento.nombre_usuario_seguro = obtener_nombre_usuario(evento.usuario, "Sin usuario registrado")
    usuario_primera_carga = primera_carga_evento.usuario if primera_carga_evento else item.subido_por
    usuario_ultima_actualizacion = ultima.usuario if ultima else item.subido_por
    return {
        "item": item,
        "eventos": eventos,
        "primera_carga": primera_carga,
        "usuario_primera_carga": usuario_primera_carga,
        "nombre_usuario_primera_carga": obtener_nombre_usuario(usuario_primera_carga),
        "ultima_actualizacion": ultima.creado_en if ultima else item.fecha_actualizacion,
        "usuario_ultima_actualizacion": usuario_ultima_actualizacion,
        "nombre_usuario_ultima_actualizacion": obtener_nombre_usuario(usuario_ultima_actualizacion),
        "decision": decision,
        "nombre_usuario_decision": (
            obtener_nombre_usuario(decision.usuario, "Sin usuario registrado") if decision else "No disponible"
        ),
        "aprobacion": aprobacion,
        "primera_revision": primera_revision,
        "cantidad_archivos": len(cargas),
        "cantidad_resubidas": len(actualizaciones),
        "cantidad_observaciones": len(observaciones),
        "ciclos_correccion": ciclos,
        "observacion_final": observaciones[-1].descripcion if observaciones else item.observacion_revision,
        "tiempo_primera_revision": formatear_duracion(primera_revision.creado_en - primera_carga) if primera_revision and primera_carga else "No disponible",
        "tiempo_total": formatear_duracion(fin - primera_carga) if primera_carga else "No disponible",
        "delta_aprobacion": aprobacion.creado_en - primera_carga if aprobacion and primera_carga else None,
        "respuestas_observaciones": respuestas,
    }


@login_required
def informe_trazabilidad_expediente_pdf(request, expediente_id):
    expediente = get_object_or_404(
        ExpedienteCAIMUS.objects.select_related(
            "asociacion", "asociacion__anio", "aprobado_por", "resolucion"
        ),
        pk=expediente_id,
    )
    if not is_admin(request.user) or not user_has_expediente_access(request.user, expediente):
        raise PermissionDenied

    historial_qs = HistorialItemExpediente.objects.select_related("usuario").order_by("creado_en", "id")
    items = list(expediente.items.filter(activo=True).select_related(
        "subido_por", "aprobado_por", "rechazado_por"
    ).prefetch_related(Prefetch("historial", queryset=historial_qs)).order_by("numero"))
    fecha_generacion = timezone.localtime(timezone.now())
    detalles = [_datos_item_trazabilidad(item, fecha_generacion) for item in items]
    primeras_cargas = [d["primera_carga"] for d in detalles if d["primera_carga"]]
    aprobaciones = [d["delta_aprobacion"] for d in detalles if d["delta_aprobacion"] is not None]
    total_archivos = sum(d["cantidad_archivos"] for d in detalles)
    # Los expedientes antiguos pueden tener archivo actual sin evento histórico.
    total_archivos += sum(1 for d in detalles if d["item"].pdf and not d["cantidad_archivos"])
    primera_carga = min(primeras_cargas) if primeras_cargas else None
    aprobado = expediente.estado == ExpedienteCAIMUS.ESTADO_APROBADO and expediente.aprobado_en
    promedio = sum(aprobaciones, timedelta()) / len(aprobaciones) if aprobaciones else None
    resolucion = getattr(expediente, "resolucion", None)
    numero_expediente = resolucion.correlativo if resolucion and resolucion.correlativo else "Sin asignar"
    if numero_expediente == "Sin asignar":
        nombre_archivo = f"Informe_Trazabilidad_Sin_Asignar_Expediente_{expediente.pk}.pdf"
    else:
        nombre_archivo = f"Informe_Trazabilidad_{numero_expediente}.pdf"
    institucion = Institucion.objects.first()
    contexto = {
        "expediente": expediente,
        "numero_expediente": numero_expediente,
        "detalles": detalles,
        "fecha_generacion": fecha_generacion,
        "usuario_generacion": request.user,
        "nombre_usuario_generacion": obtener_nombre_usuario(request.user),
        "logo_secundario_url": request.build_absolute_uri(institucion.logo2.url) if institucion and institucion.logo2 else None,
        "footer_image_url": request.build_absolute_uri(static("assets/images/pie.png")),
        "resumen": {
            "total": len(items),
            "aprobados": sum(i.estado_item == ItemChecklistCAIMUS.ESTADO_APROBADO for i in items),
            "rechazados": sum(i.estado_item == ItemChecklistCAIMUS.ESTADO_RECHAZADO for i in items),
            "pendientes": sum(i.estado_item == ItemChecklistCAIMUS.ESTADO_BORRADOR for i in items),
            "archivos": total_archivos,
            "resubidas": sum(d["cantidad_resubidas"] for d in detalles),
            "observaciones": sum(d["cantidad_observaciones"] for d in detalles),
            "primera_carga": primera_carga,
            "aprobacion_final": expediente.aprobado_en if aprobado else None,
            "tiempo_total": formatear_duracion(expediente.aprobado_en - primera_carga) if aprobado and primera_carga else "Proceso aún en curso",
            "promedio_aprobacion": formatear_duracion(promedio),
        },
    }
    html = render_to_string("asociaciones_app/informes/informe_trazabilidad_expediente.html", contexto, request=request)
    pdf = HTML(string=html, base_url=request.build_absolute_uri("/")).write_pdf()
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    return response


@asociacion_required
def resolucion_pdf(request, pk):
    expediente = get_object_or_404(ExpedienteCAIMUS, pk=pk)

    if not user_can_download_resolucion(request.user, expediente):
        raise PermissionDenied

    resolucion = getattr(expediente, "resolucion", None)

    if not is_admin(request.user) and resolucion is None:
        messages.warning(request, "La resolución aún no ha sido emitida por el administrador.")
        return redirect("asociaciones:expediente_caimus", pk=expediente.asociacion.pk)

    if is_admin(request.user) and resolucion is None:
        correlativo = generar_correlativo(expediente.asociacion.anio.anio)
        resolucion = ResolucionExpediente.objects.create(
            expediente=expediente,
            correlativo=correlativo,
            fecha_emision=timezone.now().date(),
            generado_por=request.user,
            contenido_snapshot={
                "asociacion": expediente.asociacion.nombre,
                "anio": expediente.asociacion.anio.anio,
                "estado": expediente.estado,
            },
        )

    fecha_constancia = expediente.aprobado_en or resolucion.fecha_emision or expediente.actualizado_en
    meses_es = {
        1: "enero",
        2: "febrero",
        3: "marzo",
        4: "abril",
        5: "mayo",
        6: "junio",
        7: "julio",
        8: "agosto",
        9: "septiembre",
        10: "octubre",
        11: "noviembre",
        12: "diciembre",
    }
    institucion = Institucion.objects.first()
    logo_secundario_url = None
    if institucion and institucion.logo2:
        logo_secundario_url = request.build_absolute_uri(institucion.logo2.url)
    footer_image_url = request.build_absolute_uri(static("assets/images/pie.png"))
    fecha_descarga = timezone.localtime(timezone.now())
    signer = Signer()
    codigo_validacion = signer.sign(str(expediente.id))
    validacion_url = request.build_absolute_uri(
        reverse(
            "asociaciones:validar_constancia_expediente",
            args=[codigo_validacion],
        )
    )
    qr_validacion = _generar_qr_validacion_base64(validacion_url)

    html = render_to_string(
        "asociaciones_app/resolucion_pdf.html",
        {
            "expediente": expediente,
            "resolucion": resolucion,
            "items": expediente.items.filter(activo=True).order_by("numero"),
            "fecha_constancia": fecha_constancia,
            "mes_constancia": meses_es.get(fecha_constancia.month) if fecha_constancia else "",
            "logo_secundario_url": logo_secundario_url,
            "footer_image_url": footer_image_url,
            "fecha_descarga": fecha_descarga,
            "validacion_url": validacion_url,
            "qr_validacion": qr_validacion,
        },
        request=request,
    )

    pdf = HTML(string=html, base_url=request.build_absolute_uri("/")).write_pdf()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f"inline; filename=Constancia-{resolucion.correlativo}.pdf"
    return response


def validar_constancia_expediente(request, codigo):
    signer = Signer()
    try:
        expediente_id = signer.unsign(codigo)
    except BadSignature:
        context = {
            "constancia_valida": False,
            "mensaje": "Código de validación inválido.",
            "fecha_consulta": timezone.localtime(timezone.now()),
        }
        context.update(_contexto_institucional_constancia(request))
        return render(request, "asociaciones_app/validar_constancia.html", context)

    expediente = get_object_or_404(
        ExpedienteCAIMUS.objects.select_related("asociacion", "asociacion__anio", "aprobado_por"),
        id=expediente_id,
    )

    items = expediente.items.select_related("aprobado_por", "rechazado_por").filter(activo=True).order_by("numero")
    constancia_valida = expediente.estado == ExpedienteCAIMUS.ESTADO_APROBADO

    revisores = []
    for item in items:
        if item.aprobado_por:
            revisores.append(item.aprobado_por)
        if item.rechazado_por:
            revisores.append(item.rechazado_por)

    acciones_revision = [
        HistorialItemExpediente.ACCION_APROBADO,
        HistorialItemExpediente.ACCION_RECHAZADO,
        HistorialItemExpediente.ACCION_OBSERVACION_ADMIN,
    ]
    usuarios_historial = (
        HistorialItemExpediente.objects.filter(item__expediente=expediente, accion__in=acciones_revision)
        .select_related("usuario")
        .order_by("creado_en")
    )
    for entrada in usuarios_historial:
        if entrada.usuario:
            revisores.append(entrada.usuario)

    if expediente.aprobado_por:
        revisores.append(expediente.aprobado_por)

    revisores_unicos = []
    ids = set()
    for usuario in revisores:
        if usuario and usuario.id not in ids:
            revisores_unicos.append(usuario)
            ids.add(usuario.id)

    context = {
        "expediente": expediente,
        "items": items if constancia_valida else [],
        "revisores": revisores_unicos if constancia_valida else [],
        "departamentos": DepartamentoConstancia.objects.filter(activo=True)
        .prefetch_related(
            Prefetch(
                "revisores",
                queryset=RevisorConstancia.objects.filter(activo=True)
                .select_related("usuario")
                .prefetch_related("usuario__groups")
                .order_by("orden", "usuario__first_name"),
            )
        )
        .order_by("orden", "nombre")
        if constancia_valida else [],
        "firmas": FirmaConstancia.objects.filter(activo=True).order_by("orden", "nombre") if constancia_valida else [],
        "constancia_valida": constancia_valida,
        "fecha_consulta": timezone.localtime(timezone.now()),
        "codigo_validacion": codigo,
    }
    context.update(_contexto_institucional_constancia(request))
    return render(request, "asociaciones_app/validar_constancia.html", context)


def informatica_required(view_func):
    @login_required
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.groups.filter(name__iexact="Informatica").exists():
            raise PermissionDenied
        return view_func(request, *args, **kwargs)

    return _wrapped


@informatica_required
def firmas_constancia_list(request):
    firmas = FirmaConstancia.objects.all().order_by("orden", "nombre")
    return render(request, "asociaciones/firmas_constancia/list.html", {"firmas": firmas})


@informatica_required
def firma_constancia_create(request):
    if request.method == "POST":
        form = FirmaConstanciaForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Firma creada correctamente.")
            return redirect("asociaciones:firmas_constancia_list")
    else:
        form = FirmaConstanciaForm()

    return render(
        request,
        "asociaciones/firmas_constancia/form.html",
        {"form": form, "titulo": "Nueva firma de constancia", "accion": "Crear"},
    )


@informatica_required
def firma_constancia_update(request, pk):
    firma = get_object_or_404(FirmaConstancia, pk=pk)
    if request.method == "POST":
        form = FirmaConstanciaForm(request.POST, request.FILES, instance=firma)
        if form.is_valid():
            form.save()
            messages.success(request, "Firma actualizada correctamente.")
            return redirect("asociaciones:firmas_constancia_list")
    else:
        form = FirmaConstanciaForm(instance=firma)

    return render(
        request,
        "asociaciones/firmas_constancia/form.html",
        {"form": form, "firma": firma, "titulo": "Editar firma de constancia", "accion": "Actualizar"},
    )


@informatica_required
def firma_constancia_toggle(request, pk):
    firma = get_object_or_404(FirmaConstancia, pk=pk)
    firma.activo = not firma.activo
    firma.save(update_fields=["activo"])
    messages.success(request, "Estado de la firma actualizado correctamente.")
    return redirect("asociaciones:firmas_constancia_list")


@informatica_required
def departamentos_constancia_list(request):
    departamentos = DepartamentoConstancia.objects.order_by("orden", "nombre")
    return render(request, "asociaciones/departamentos_constancia/list.html", {"departamentos": departamentos})


@informatica_required
def departamento_constancia_create(request):
    if request.method == "POST":
        form = DepartamentoConstanciaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Departamento creado correctamente.")
            return redirect("asociaciones:departamentos_constancia_list")
    else:
        form = DepartamentoConstanciaForm()

    return render(
        request,
        "asociaciones/departamentos_constancia/form.html",
        {"form": form, "titulo": "Nuevo departamento de constancia", "accion": "Crear"},
    )


@informatica_required
def departamento_constancia_update(request, pk):
    departamento = get_object_or_404(DepartamentoConstancia, pk=pk)
    if request.method == "POST":
        form = DepartamentoConstanciaForm(request.POST, instance=departamento)
        if form.is_valid():
            form.save()
            messages.success(request, "Departamento actualizado correctamente.")
            return redirect("asociaciones:departamentos_constancia_list")
    else:
        form = DepartamentoConstanciaForm(instance=departamento)

    return render(
        request,
        "asociaciones/departamentos_constancia/form.html",
        {"form": form, "titulo": "Editar departamento de constancia", "accion": "Actualizar"},
    )


@informatica_required
def departamento_constancia_toggle(request, pk):
    departamento = get_object_or_404(DepartamentoConstancia, pk=pk)
    departamento.activo = not departamento.activo
    departamento.save(update_fields=["activo"])
    messages.success(request, "Estado del departamento actualizado correctamente.")
    return redirect("asociaciones:departamentos_constancia_list")


@informatica_required
def revisores_constancia_list(request):
    revisores = RevisorConstancia.objects.select_related("departamento", "usuario").prefetch_related("usuario__groups").order_by(
        "departamento__orden", "orden", "usuario__first_name"
    )
    return render(request, "asociaciones/revisores_constancia/list.html", {"revisores": revisores})


@informatica_required
def revisor_constancia_create(request):
    if request.method == "POST":
        form = RevisorConstanciaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Revisor creado correctamente.")
            return redirect("asociaciones:revisores_constancia_list")
    else:
        form = RevisorConstanciaForm()

    return render(
        request,
        "asociaciones/revisores_constancia/form.html",
        {"form": form, "titulo": "Nuevo revisor de constancia", "accion": "Crear"},
    )


@informatica_required
def revisor_constancia_update(request, pk):
    revisor = get_object_or_404(RevisorConstancia, pk=pk)
    if request.method == "POST":
        form = RevisorConstanciaForm(request.POST, instance=revisor)
        if form.is_valid():
            form.save()
            messages.success(request, "Revisor actualizado correctamente.")
            return redirect("asociaciones:revisores_constancia_list")
    else:
        form = RevisorConstanciaForm(instance=revisor)

    return render(
        request,
        "asociaciones/revisores_constancia/form.html",
        {"form": form, "titulo": "Editar revisor de constancia", "accion": "Actualizar"},
    )


@informatica_required
def revisor_constancia_toggle(request, pk):
    revisor = get_object_or_404(RevisorConstancia, pk=pk)
    revisor.activo = not revisor.activo
    revisor.save(update_fields=["activo"])
    messages.success(request, "Estado del revisor actualizado correctamente.")
    return redirect("asociaciones:revisores_constancia_list")


firma_constancia_edit = firma_constancia_update


@informatica_required
@require_POST
def firma_constancia_delete(request, pk):
    firma = get_object_or_404(FirmaConstancia, pk=pk)
    firma.delete()
    messages.success(request, "Firma eliminada correctamente.")
    return redirect("asociaciones:firmas_constancia_list")


@login_required
@asociacion_required
@require_POST
def notificaciones_marcar_leidas(request, asociacion_id):
    asociacion = get_object_or_404(Asociacion, pk=asociacion_id)
    if not user_has_asociacion_access(request.user, asociacion):
        raise PermissionDenied
    asociacion.notificaciones.filter(leida=False).update(leida=True)
    messages.success(request, "Alertas marcadas como leídas.")
    return redirect("asociaciones:mis_asociaciones")
    EntradaRevisionAdmin,
